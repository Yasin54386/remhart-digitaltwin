"""
REMHART Digital Twin - Reports Router
=======================================
Handles report generation and data export functionality.

Endpoints:
- GET /api/reports/preview - Preview filtered data
- GET /api/reports/export/excel - Export data to Excel
- GET /api/reports/export/csv - Export data to CSV

Author: REMHART Team
Date: 2025
"""

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from typing import Optional, List
from datetime import datetime, timedelta
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models.db_models import DateTimeTable
from app.utils.security import get_current_user

router = APIRouter()


def build_query(
    db: Session,
    start_time: Optional[datetime],
    end_time: Optional[datetime],
    is_simulation: Optional[bool],
    simulation_id: Optional[str],
    limit: Optional[int] = None
):
    """Build database query with filters"""
    query = db.query(DateTimeTable).order_by(desc(DateTimeTable.timestamp))

    # Apply time filters
    if start_time:
        query = query.filter(DateTimeTable.timestamp >= start_time)
    if end_time:
        query = query.filter(DateTimeTable.timestamp <= end_time)

    # Apply simulation filters
    if is_simulation is not None:
        query = query.filter(DateTimeTable.is_simulation == is_simulation)
    if simulation_id:
        query = query.filter(DateTimeTable.simulation_id == simulation_id)

    # Apply limit if specified
    if limit:
        query = query.limit(limit)

    return query


def format_data_record(dt_record):
    """Format a database record into a dictionary"""
    return {
        "timestamp": dt_record.timestamp.isoformat(),
        "voltage_a": dt_record.voltage[0].phaseA if dt_record.voltage else 0,
        "voltage_b": dt_record.voltage[0].phaseB if dt_record.voltage else 0,
        "voltage_c": dt_record.voltage[0].phaseC if dt_record.voltage else 0,
        "voltage_avg": dt_record.voltage[0].average if dt_record.voltage else 0,
        "current_a": dt_record.current[0].phaseA if dt_record.current else 0,
        "current_b": dt_record.current[0].phaseB if dt_record.current else 0,
        "current_c": dt_record.current[0].phaseC if dt_record.current else 0,
        "current_avg": dt_record.current[0].average if dt_record.current else 0,
        "frequency": dt_record.frequency[0].frequency_value if dt_record.frequency else 50.0,
        "active_power_a": dt_record.active_power[0].phaseA if dt_record.active_power else 0,
        "active_power_b": dt_record.active_power[0].phaseB if dt_record.active_power else 0,
        "active_power_c": dt_record.active_power[0].phaseC if dt_record.active_power else 0,
        "active_power_total": dt_record.active_power[0].total if dt_record.active_power else 0,
        "reactive_power_a": dt_record.reactive_power[0].phaseA if dt_record.reactive_power else 0,
        "reactive_power_b": dt_record.reactive_power[0].phaseB if dt_record.reactive_power else 0,
        "reactive_power_c": dt_record.reactive_power[0].phaseC if dt_record.reactive_power else 0,
        "reactive_power_total": dt_record.reactive_power[0].total if dt_record.reactive_power else 0,
        "is_simulation": dt_record.is_simulation if hasattr(dt_record, 'is_simulation') else False,
        "simulation_name": dt_record.simulation_name if hasattr(dt_record, 'simulation_name') else None,
    }


@router.get("/preview")
async def preview_report(
    start_time: Optional[datetime] = Query(default=None),
    end_time: Optional[datetime] = Query(default=None),
    is_simulation: Optional[bool] = Query(default=None),
    simulation_id: Optional[str] = Query(default=None),
    limit: int = Query(default=100, ge=1, le=1000),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Preview filtered data before export.
    Returns limited dataset for preview purposes.
    """
    try:
        query = build_query(db, start_time, end_time, is_simulation, simulation_id, limit)
        records = query.all()

        # Format records
        data = [format_data_record(record) for record in records]

        # Calculate summary statistics
        if data:
            voltage_values = [d['voltage_avg'] for d in data]
            current_values = [d['current_avg'] for d in data]
            power_values = [d['active_power_total'] for d in data]

            summary = {
                "total_records": len(data),
                "voltage_avg": round(sum(voltage_values) / len(voltage_values), 2),
                "voltage_min": round(min(voltage_values), 2),
                "voltage_max": round(max(voltage_values), 2),
                "current_avg": round(sum(current_values) / len(current_values), 2),
                "current_min": round(min(current_values), 2),
                "current_max": round(max(current_values), 2),
                "power_avg": round(sum(power_values) / len(power_values), 2),
                "power_min": round(min(power_values), 2),
                "power_max": round(max(power_values), 2),
            }
        else:
            summary = {"total_records": 0}

        return {
            "success": True,
            "summary": summary,
            "data": data,
            "filters": {
                "start_time": start_time.isoformat() if start_time else None,
                "end_time": end_time.isoformat() if end_time else None,
                "is_simulation": is_simulation,
                "simulation_id": simulation_id
            }
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error previewing report: {str(e)}")


@router.get("/export/csv")
async def export_csv(
    start_time: Optional[datetime] = Query(default=None),
    end_time: Optional[datetime] = Query(default=None),
    is_simulation: Optional[bool] = Query(default=None),
    simulation_id: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Export filtered data to CSV format.
    """
    try:
        query = build_query(db, start_time, end_time, is_simulation, simulation_id)
        records = query.all()

        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow([
            'Timestamp',
            'Voltage A (V)', 'Voltage B (V)', 'Voltage C (V)', 'Voltage Avg (V)',
            'Current A (A)', 'Current B (A)', 'Current C (A)', 'Current Avg (A)',
            'Frequency (Hz)',
            'Active Power A (W)', 'Active Power B (W)', 'Active Power C (W)', 'Active Power Total (W)',
            'Reactive Power A (VAR)', 'Reactive Power B (VAR)', 'Reactive Power C (VAR)', 'Reactive Power Total (VAR)',
            'Is Simulation', 'Simulation Name'
        ])

        # Write data rows
        for record in records:
            data = format_data_record(record)
            writer.writerow([
                data['timestamp'],
                data['voltage_a'], data['voltage_b'], data['voltage_c'], data['voltage_avg'],
                data['current_a'], data['current_b'], data['current_c'], data['current_avg'],
                data['frequency'],
                data['active_power_a'], data['active_power_b'], data['active_power_c'], data['active_power_total'],
                data['reactive_power_a'], data['reactive_power_b'], data['reactive_power_c'], data['reactive_power_total'],
                data['is_simulation'], data['simulation_name'] or ''
            ])

        # Prepare response
        output.seek(0)
        filename = f"grid_data_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating CSV: {str(e)}")


@router.get("/export/excel")
async def export_excel(
    start_time: Optional[datetime] = Query(default=None),
    end_time: Optional[datetime] = Query(default=None),
    is_simulation: Optional[bool] = Query(default=None),
    simulation_id: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Export filtered data to Excel format with formatting and summary.
    """
    try:
        query = build_query(db, start_time, end_time, is_simulation, simulation_id)
        records = query.all()

        # Create workbook
        wb = Workbook()

        # --- SUMMARY SHEET ---
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Title styling
        title_font = Font(size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")

        # Add title
        ws_summary['A1'] = "REMHART Grid Data Report"
        ws_summary['A1'].font = title_font
        ws_summary['A1'].fill = title_fill
        ws_summary.merge_cells('A1:B1')

        # Report metadata
        row = 3
        ws_summary[f'A{row}'] = "Generated:"
        ws_summary[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 1
        ws_summary[f'A{row}'] = "Generated by:"
        ws_summary[f'B{row}'] = current_user.get('email', current_user.get('sub', 'Unknown'))
        row += 1
        ws_summary[f'A{row}'] = "Time Range:"
        time_range = f"{start_time.strftime('%Y-%m-%d %H:%M') if start_time else 'All'} to {end_time.strftime('%Y-%m-%d %H:%M') if end_time else 'All'}"
        ws_summary[f'B{row}'] = time_range
        row += 1
        ws_summary[f'A{row}'] = "Data Source:"
        if is_simulation is None:
            source = "All Data"
        elif is_simulation:
            source = f"Simulation ({simulation_id if simulation_id else 'All'})"
        else:
            source = "Real-time"
        ws_summary[f'B{row}'] = source
        row += 1
        ws_summary[f'A{row}'] = "Total Records:"
        ws_summary[f'B{row}'] = len(records)

        # Calculate statistics
        if records:
            data_list = [format_data_record(r) for r in records]
            voltage_values = [d['voltage_avg'] for d in data_list]
            current_values = [d['current_avg'] for d in data_list]
            power_values = [d['active_power_total'] for d in data_list]

            row += 2
            ws_summary[f'A{row}'] = "Statistics"
            ws_summary[f'A{row}'].font = header_font
            ws_summary[f'A{row}'].fill = header_fill
            ws_summary.merge_cells(f'A{row}:D{row}')

            row += 1
            ws_summary[f'A{row}'] = "Parameter"
            ws_summary[f'B{row}'] = "Average"
            ws_summary[f'C{row}'] = "Minimum"
            ws_summary[f'D{row}'] = "Maximum"
            for col in ['A', 'B', 'C', 'D']:
                ws_summary[f'{col}{row}'].font = Font(bold=True)

            row += 1
            ws_summary[f'A{row}'] = "Voltage (V)"
            ws_summary[f'B{row}'] = round(sum(voltage_values) / len(voltage_values), 2)
            ws_summary[f'C{row}'] = round(min(voltage_values), 2)
            ws_summary[f'D{row}'] = round(max(voltage_values), 2)

            row += 1
            ws_summary[f'A{row}'] = "Current (A)"
            ws_summary[f'B{row}'] = round(sum(current_values) / len(current_values), 2)
            ws_summary[f'C{row}'] = round(min(current_values), 2)
            ws_summary[f'D{row}'] = round(max(current_values), 2)

            row += 1
            ws_summary[f'A{row}'] = "Power (W)"
            ws_summary[f'B{row}'] = round(sum(power_values) / len(power_values), 2)
            ws_summary[f'C{row}'] = round(min(power_values), 2)
            ws_summary[f'D{row}'] = round(max(power_values), 2)

        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 25
        ws_summary.column_dimensions['C'].width = 15
        ws_summary.column_dimensions['D'].width = 15

        # --- DATA SHEET ---
        ws_data = wb.create_sheet("Data")

        # Headers
        headers = [
            'Timestamp',
            'Voltage A', 'Voltage B', 'Voltage C', 'Voltage Avg',
            'Current A', 'Current B', 'Current C', 'Current Avg',
            'Frequency',
            'Power A', 'Power B', 'Power C', 'Power Total',
            'Is Simulation', 'Simulation'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for row_num, record in enumerate(records, 2):
            data = format_data_record(record)
            ws_data.cell(row=row_num, column=1, value=data['timestamp'])
            ws_data.cell(row=row_num, column=2, value=data['voltage_a'])
            ws_data.cell(row=row_num, column=3, value=data['voltage_b'])
            ws_data.cell(row=row_num, column=4, value=data['voltage_c'])
            ws_data.cell(row=row_num, column=5, value=data['voltage_avg'])
            ws_data.cell(row=row_num, column=6, value=data['current_a'])
            ws_data.cell(row=row_num, column=7, value=data['current_b'])
            ws_data.cell(row=row_num, column=8, value=data['current_c'])
            ws_data.cell(row=row_num, column=9, value=data['current_avg'])
            ws_data.cell(row=row_num, column=10, value=data['frequency'])
            ws_data.cell(row=row_num, column=11, value=data['active_power_a'])
            ws_data.cell(row=row_num, column=12, value=data['active_power_b'])
            ws_data.cell(row=row_num, column=13, value=data['active_power_c'])
            ws_data.cell(row=row_num, column=14, value=data['active_power_total'])
            ws_data.cell(row=row_num, column=15, value='Yes' if data['is_simulation'] else 'No')
            ws_data.cell(row=row_num, column=16, value=data['simulation_name'] or '')

        # Auto-adjust column widths
        for column in ws_data.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_data.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws_data.freeze_panes = 'A2'

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"grid_data_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {str(e)}")
